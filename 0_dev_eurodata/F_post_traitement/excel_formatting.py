# -*- coding: utf-8 -*-
"""
Created on Thu Feb  8 13:15:44 2024
@author: Andry-Harifetra

this script :
    1) add a blank row on header
    2) copy/paste value (by model file)
    3) fill bg, setting up style
    * run entire code ONCE *
"""

import openpyxl
from openpyxl.styles import NamedStyle,Alignment, Border, Side, Font

def copy_value(plage_source,plage_destination):
    for row_source, row_destination in zip(plage_source, plage_destination):
        for cell_source, cell_destination in zip(row_source, row_destination):
            cell_destination.value = cell_source.value
            #print(f"src : {cell_source.value}")

def fill_range(target_range, color):
    # fill background of a range
    for row in target_range:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')

def wrap_range(target_range,workbook):
    # Set style for a given range
    style = openpyxl.styles.NamedStyle(
        name='center_style_5',
        alignment=Alignment(horizontal='center', vertical='center',wrap_text=True),
        border = Border(
                left=Side(border_style='thin', color='0E0E0E'),    # Bordure à gauche
                right=Side(border_style='thin', color='0E0E0E'),   # Bordure à droite
                top=Side(border_style='thin', color='0E0E0E'),     # Bordure en haut
                bottom=Side(border_style='thin', color='0E0E0E')   # Bordure en bas
                ),
        font=Font(bold=True))
    workbook.add_named_style(style)
    for row in target_range:
        for cell in row:
            cell.style = 'center_style_5'
###################################################################################

# get region list
import pandas as pd
from G_GLOBAL.variable import PROJECT_PATH, REMOTE_PATH

def run_excel_formating():
    input_path = PROJECT_PATH+r'B_data_processing/OUTPUT/'
    # EXPLICIT OUTPUT PATH
    #ouput_path = PROJECT_PATH+r'F_post_traitement/OUTPUT/'
    ouput_path = REMOTE_PATH
    model_path = PROJECT_PATH+r'F_post_traitement/INPUT/model.xlsx'
    
    
    
    df_prevalid_cds = pd.read_excel(input_path+'cds_prevalidation.xlsx')
    region_list = df_prevalid_cds[df_prevalid_cds['Statut']=="En cours"]['Region'].value_counts().reset_index()['index']
    
    # Mise en forme
    for region in region_list :
        file_path = ouput_path+region+".xlsx"
        
        workbook = openpyxl.load_workbook(file_path)
        workbook.active.insert_rows(1) # * insert a head row once
            # HEADER VALUE     
        source = openpyxl.load_workbook(model_path)
        copy_value(source.active['A1:V2'],workbook.active['A1:V2'])
    
            # HEADER STYLE
        wrap_range(workbook.active['A2:V2'],workbook)
    
            # CELL FILL
        fill_range(workbook.active['H1:L2'],color="CCC0DA")
        fill_range(workbook.active['M1:O2'],color="E6B8B7")
        fill_range(workbook.active['P2:R2'],color="B8CCE4")
        fill_range(workbook.active['T1:V2'],color="E4DFEC")
    
        green = "D8E4BC"
        fill_range(workbook.active['H3:I500'],color=green)
        fill_range(workbook.active['O3:P500'],color=green)
        fill_range(workbook.active['R3:R500'],color=green)
        fill_range(workbook.active['L3:L500'],color=green)
        
        orange = "FDE9D9"
        fill_range(workbook.active['J3:J500'],color=orange)
        fill_range(workbook.active['Q3:Q500'],color=orange)
        
        # set column width
        workbook.active.column_dimensions['B'].width = 25
        workbook.active.column_dimensions['C'].width = 15
        workbook.active.column_dimensions['D'].width = 10
        
        workbook.save(file_path)
    
    print("formating done !")

if __name__=="__main__":
    run_excel_formating()